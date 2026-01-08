import datetime
import json
import requests
import os
import io
import logging
import traceback
import base64
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import azure.functions as func

# Setup logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

def get_access_token():
    """Get Azure access token using service principal credentials"""
    try:
        logger.info("Starting token acquisition...")
        
        TENANT_ID = os.environ.get("TENANT_ID")
        CLIENT_ID = os.environ.get("CLIENT_ID")
        CLIENT_SECRET = os.environ.get("CLIENT_SECRET")
        
        # Detailed validation
        if not TENANT_ID:
            raise ValueError("TENANT_ID environment variable is not set")
        if not CLIENT_ID:
            raise ValueError("CLIENT_ID environment variable is not set")
        if not CLIENT_SECRET:
            raise ValueError("CLIENT_SECRET environment variable is not set")
        
        logger.info(f"TENANT_ID: {TENANT_ID[:8]}... (length: {len(TENANT_ID)})")
        logger.info(f"CLIENT_ID: {CLIENT_ID[:8]}... (length: {len(CLIENT_ID)})")
        logger.info(f"CLIENT_SECRET: {'*' * 8}... (length: {len(CLIENT_SECRET)})")
        
        url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/token"
        payload = {
            "grant_type": "client_credentials",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "resource": "https://management.azure.com/"
        }
        
        logger.info(f"Requesting token from: {url}")
        response = requests.post(url, data=payload, timeout=30)
        
        if response.status_code != 200:
            logger.error(f"Token request failed with status {response.status_code}")
            logger.error(f"Response: {response.text}")
            response.raise_for_status()
        
        token_data = response.json()
        logger.info("Access token acquired successfully")
        return token_data["access_token"]
        
    except requests.exceptions.Timeout as e:
        logger.error(f"Timeout while getting access token: {str(e)}")
        raise Exception(f"Authentication timeout: {str(e)}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Request error getting access token: {str(e)}")
        if hasattr(e.response, 'text'):
            logger.error(f"Error response: {e.response.text}")
        raise Exception(f"Authentication failed: {str(e)}")
    except KeyError as e:
        logger.error(f"Missing key in token response: {str(e)}")
        raise Exception(f"Invalid token response: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error getting access token: {str(e)}")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise

def get_previous_month_range():
    """Calculate the first and last day of the previous month"""
    try:
        today = datetime.date.today()
        first_day_this_month = today.replace(day=1)
        last_day_prev_month = first_day_this_month - datetime.timedelta(days=1)
        first_day_prev_month = last_day_prev_month.replace(day=1)
        
        start_date = first_day_prev_month.isoformat()
        end_date = last_day_prev_month.isoformat()
        
        logger.info(f"Date range calculated: {start_date} to {end_date}")
        return start_date, end_date
        
    except Exception as e:
        logger.error(f"Error calculating date range: {str(e)}")
        raise

def get_all_subscriptions(token):
    """Fetch all subscriptions accessible to the service principal"""
    try:
        logger.info("Fetching subscriptions...")
        url = "https://management.azure.com/subscriptions?api-version=2020-01-01"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        response = requests.get(url, headers=headers, timeout=30)
        
        if response.status_code != 200:
            logger.error(f"Subscription fetch failed with status {response.status_code}")
            logger.error(f"Response: {response.text}")
            response.raise_for_status()
        
        subscriptions = response.json().get("value", [])
        logger.info(f"Found {len(subscriptions)} subscriptions")
        
        if not subscriptions:
            logger.warning("No subscriptions found for this service principal")
        else:
            for sub in subscriptions[:3]:  # Log first 3 subscriptions
                logger.info(f"  - {sub.get('displayName')} ({sub.get('subscriptionId')})")
        
        return subscriptions
        
    except requests.exceptions.Timeout as e:
        logger.error(f"Timeout fetching subscriptions: {str(e)}")
        raise Exception(f"Subscription fetch timeout: {str(e)}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching subscriptions: {str(e)}")
        if hasattr(e.response, 'text'):
            logger.error(f"Error response: {e.response.text}")
        raise Exception(f"Failed to fetch subscriptions: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error fetching subscriptions: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise

def fetch_cost_for_subscription(token, subscription_id, start_date, end_date):
    """Fetch cost data for a specific subscription"""
    try:
        logger.info(f"Fetching cost for subscription: {subscription_id}")
        
        url = f"https://management.azure.com/subscriptions/{subscription_id}/providers/Microsoft.CostManagement/query?api-version=2023-03-01"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        body = {
            "type": "ActualCost",
            "timeframe": "Custom",
            "timePeriod": {
                "from": start_date,
                "to": end_date
            },
            "dataset": {
                "granularity": "None",
                "aggregation": {
                    "totalCost": {
                        "name": "Cost",
                        "function": "Sum"
                    }
                }
            }
        }
        
        response = requests.post(url, headers=headers, json=body, timeout=60)
        
        if response.status_code != 200:
            logger.warning(f"Cost fetch failed for {subscription_id} with status {response.status_code}")
            logger.warning(f"Response: {response.text}")
            # Return empty structure instead of raising error
            return {"properties": {"rows": [], "columns": []}}
        
        cost_data = response.json()
        rows = cost_data.get("properties", {}).get("rows", [])
        logger.info(f"  Cost data retrieved: {len(rows)} rows")
        
        return cost_data
        
    except requests.exceptions.Timeout as e:
        logger.warning(f"Timeout fetching cost for {subscription_id}: {str(e)}")
        return {"properties": {"rows": [], "columns": []}}
    except Exception as e:
        logger.warning(f"Error fetching cost for {subscription_id}: {str(e)}")
        return {"properties": {"rows": [], "columns": []}}

def generate_excel(all_costs_data, start_date, end_date):
    """Generate Excel with all subscriptions cost data"""
    try:
        logger.info("Generating Excel file...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Azure Cost Report"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Add headers
        headers = ["Subscription Name", "Subscription ID", "From Date", "To Date", "Total Cost (USD)", "Status"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        total_cost_all = 0.0
        
        # Add data for each subscription
        for sub_data in all_costs_data:
            subscription_name = sub_data["subscription_name"]
            subscription_id = sub_data["subscription_id"]
            cost_data = sub_data["cost_data"]
            
            # Extract cost
            rows = cost_data.get("properties", {}).get("rows", [])
            if not rows or len(rows) == 0:
                total_cost = 0.0
                status = "No usage data"
            else:
                total_cost = float(rows[0][0]) if len(rows[0]) > 0 else 0.0
                status = "Active" if total_cost > 0 else "No charges"
            
            total_cost_all += total_cost
            
            ws.append([
                subscription_name,
                subscription_id,
                start_date,
                end_date,
                round(total_cost, 2),
                status
            ])
        
        # Add total row
        ws.append([])
        total_row = ws.max_row
        ws.append(["TOTAL", "", "", "", round(total_cost_all, 2), ""])
        
        # Style total row
        for cell in ws[total_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        # Add summary info
        ws.append([])
        ws.append([f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"Total Subscriptions: {len(all_costs_data)}"])
        
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        logger.info(f"Excel file generated successfully ({len(all_costs_data)} subscriptions, Total: ${round(total_cost_all, 2)})")
        return file_stream, total_cost_all
        
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise

def send_email_with_attachment(excel_file, filename, start_date, end_date, total_cost, subscription_count):
    """Send email with Excel attachment using Gmail SMTP"""
    try:
        logger.info("Preparing to send email via Gmail SMTP...")
        
        # Get Gmail credentials from environment variables
        GMAIL_USER = os.environ.get("GMAIL_USER")
        GMAIL_PASSWORD = os.environ.get("GMAIL_PASSWORD")
        TO_EMAIL = "vardhanullas7@gmail.com"
        
        if not GMAIL_USER:
            raise ValueError("GMAIL_USER environment variable is not set")
        if not GMAIL_PASSWORD:
            raise ValueError("GMAIL_PASSWORD environment variable is not set (use App Password)")
        
        logger.info(f"Sending from: {GMAIL_USER}")
        logger.info(f"Sending to: {TO_EMAIL}")
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = GMAIL_USER
        msg['To'] = TO_EMAIL
        msg['Subject'] = f"Azure Cost Report - {start_date} to {end_date}"
        
        # HTML email body
        html_body = f"""
        <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #0078D4; border-bottom: 3px solid #0078D4; padding-bottom: 10px;">
                        Azure Cost Report
                    </h2>
                    
                    <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <h3 style="margin-top: 0; color: #0078D4;">Report Summary</h3>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr>
                                <td style="padding: 8px 0;"><strong>Period:</strong></td>
                                <td style="padding: 8px 0;">{start_date} to {end_date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0;"><strong>Total Subscriptions:</strong></td>
                                <td style="padding: 8px 0;">{subscription_count}</td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0;"><strong>Total Cost:</strong></td>
                                <td style="padding: 8px 0; color: #0078D4; font-size: 18px;">
                                    <strong>${total_cost:.2f} USD</strong>
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 8px 0;"><strong>Generated:</strong></td>
                                <td style="padding: 8px 0;">{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <p>Please find the detailed Azure cost report attached to this email.</p>
                    
                    <p>The Excel file contains:</p>
                    <ul>
                        <li>Individual subscription costs</li>
                        <li>Subscription status</li>
                        <li>Date range details</li>
                        <li>Total cost summary</li>
                    </ul>
                    
                    <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 12px; color: #666;">
                        <p>This is an automated report generated by Azure Cost Management Function.</p>
                        <p>If you have any questions, please contact your Azure administrator.</p>
                    </div>
                </div>
            </body>
        </html>
        """
        
        # Attach HTML body
        msg.attach(MIMEText(html_body, 'html'))
        
        # Attach Excel file
        excel_file.seek(0)
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(excel_file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)
        
        # Send email via Gmail SMTP
        logger.info("Connecting to Gmail SMTP server...")
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        
        logger.info("Logging in to Gmail...")
        server.login(GMAIL_USER, GMAIL_PASSWORD)
        
        logger.info("Sending email...")
        text = msg.as_string()
        server.sendmail(GMAIL_USER, TO_EMAIL, text)
        server.quit()
        
        logger.info(f"✓ Email sent successfully to: {TO_EMAIL}")
        return True
        
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"Gmail authentication failed: {str(e)}")
        logger.error("Make sure you're using an App Password, not your regular Gmail password")
        logger.error("Generate App Password at: https://myaccount.google.com/apppasswords")
        raise
    except Exception as e:
        logger.error(f"Error sending email: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise

def main(req: func.HttpRequest) -> func.HttpResponse:
    """Main function entry point"""
    logger.info('=' * 80)
    logger.info('Azure Cost Report Function - Starting execution')
    logger.info('=' * 80)
    
    try:
        # Step 1: Validate environment variables
        logger.info("Step 1: Validating environment variables...")
        required_vars = ["TENANT_ID", "CLIENT_ID", "CLIENT_SECRET", "GMAIL_USER", "GMAIL_PASSWORD"]
        missing_vars = [var for var in required_vars if not os.environ.get(var)]
        
        if missing_vars:
            error_msg = f"Missing environment variables: {', '.join(missing_vars)}"
            logger.error(error_msg)
            logger.error("Please configure these in Azure Function App Settings")
            return func.HttpResponse(
                body=json.dumps({
                    "error": error_msg,
                    "details": "Configure environment variables in Azure Portal → Function App → Configuration → Application Settings"
                }),
                status_code=500,
                mimetype="application/json"
            )
        
        logger.info("✓ All environment variables present")
        
        # Step 2: Get access token
        logger.info("Step 2: Acquiring Azure access token...")
        token = get_access_token()
        logger.info("✓ Access token acquired")
        
        # Step 3: Calculate date range
        logger.info("Step 3: Calculating date range...")
        start_date, end_date = get_previous_month_range()
        logger.info(f"✓ Date range: {start_date} to {end_date}")
        
        # Step 4: Fetch subscriptions
        logger.info("Step 4: Fetching all subscriptions...")
        subscriptions = get_all_subscriptions(token)
        
        if not subscriptions:
            logger.warning("No subscriptions found")
            return func.HttpResponse(
                body=json.dumps({
                    "error": "No subscriptions found",
                    "details": "The service principal has no access to any subscriptions"
                }),
                status_code=404,
                mimetype="application/json"
            )
        
        logger.info(f"✓ Found {len(subscriptions)} subscriptions")
        
        # Step 5: Fetch cost data for each subscription
        logger.info("Step 5: Fetching cost data for all subscriptions...")
        all_costs_data = []
        
        for idx, subscription in enumerate(subscriptions, 1):
            sub_id = subscription.get("subscriptionId")
            sub_name = subscription.get("displayName", "Unknown")
            
            logger.info(f"  [{idx}/{len(subscriptions)}] Processing: {sub_name}")
            
            cost_data = fetch_cost_for_subscription(token, sub_id, start_date, end_date)
            
            all_costs_data.append({
                "subscription_id": sub_id,
                "subscription_name": sub_name,
                "cost_data": cost_data
            })
        
        logger.info("✓ Cost data fetched for all subscriptions")
        
        # Step 6: Generate Excel file
        logger.info("Step 6: Generating Excel report...")
        excel_file, total_cost = generate_excel(all_costs_data, start_date, end_date)
        logger.info("✓ Excel report generated")
        
        # Step 7: Send email
        filename = f"azure_all_subscriptions_cost_{start_date}_to_{end_date}.xlsx"
        logger.info(f"Step 7: Sending email with attachment: {filename}")
        
        send_email_with_attachment(
            excel_file, 
            filename, 
            start_date, 
            end_date, 
            total_cost,
            len(all_costs_data)
        )
        
        logger.info("✓ Email sent successfully")
        logger.info('=' * 80)
        logger.info('Execution completed successfully!')
        logger.info('=' * 80)
        
        return func.HttpResponse(
            body=json.dumps({
                "status": "success",
                "message": f"Cost report email sent successfully to vardhanullas7@gmail.com",
                "report_period": f"{start_date} to {end_date}",
                "total_subscriptions": len(all_costs_data),
                "total_cost": round(total_cost, 2),
                "filename": filename
            }),
            status_code=200,
            mimetype="application/json"
        )
        
    except ValueError as ve:
        error_msg = f"Configuration error: {str(ve)}"
        logger.error('=' * 80)
        logger.error('EXECUTION FAILED - Configuration Error')
        logger.error('=' * 80)
        logger.error(error_msg)
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        return func.HttpResponse(
            body=json.dumps({
                "error": error_msg,
                "type": "ConfigurationError",
                "traceback": traceback.format_exc()
            }),
            status_code=500,
            mimetype="application/json"
        )
        
    except requests.exceptions.RequestException as re:
        error_msg = f"Azure API error: {str(re)}"
        logger.error('=' * 80)
        logger.error('EXECUTION FAILED - API Error')
        logger.error('=' * 80)
        logger.error(error_msg)
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        return func.HttpResponse(
            body=json.dumps({
                "error": error_msg,
                "type": "APIError",
                "traceback": traceback.format_exc()
            }),
            status_code=500,
            mimetype="application/json"
        )
        
    except Exception as e:
        error_msg = f"Unexpected error: {str(e)}"
        logger.error('=' * 80)
        logger.error('EXECUTION FAILED - Unexpected Error')
        logger.error('=' * 80)
        logger.error(error_msg)
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        return func.HttpResponse(
            body=json.dumps({
                "error": str(e),
                "type": type(e).__name__,
                "traceback": traceback.format_exc()
            }),
            status_code=500,
            mimetype="application/json"
        )